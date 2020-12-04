import gensim
from gensim.models import Word2Vec, KeyedVectors
import numpy as np
import nltk
import pandas as pd
import itertools
from nltk.corpus import stopwords
from nltk.tokenize import sent_tokenize, word_tokenize
import scipy
from scipy import spatial
from nltk.tokenize.toktok import ToktokTokenizer
import re

import openpyxl


tokenizer = ToktokTokenizer()
stopword_list = nltk.corpus.stopwords.words('english')

class NlpAlgo():


    def __init__(self):
        return self.model=model



    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb['Sheet1']
    tempList = []
    fin= []
    #iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row)):
    for row in ws['A{}:A{}'.format(ws.min_row,ws.max_row)]:             
        for cell in row:
            if cell != None:
                tempList.append(cell.value)

    for var in tempList:
        if var != None:
            fin.append(var)

    

    EMBEDDING_FILE = "obs_Word_Embeddings.bin"
    model = gensim.models.KeyedVectors.load_word2vec_format(EMBEDDING_FILE, binary= True)
    #model = Word2Vec.load("/content/word2vec_Model.bin")

    def remove_stopwords(self,text, is_lower_case=False):
        pattern = r'[^a-zA-z0-9\s]' 
        text1 = re.sub(pattern, ", ", ''.join(text))
        tokens = tokenizer.tokenize(text1)
        tokens = [token.strip() for token in tokens]
        if is_lower_case:
            filtered_tokens = [token for token in tokens if token not in stopword_list]
        else:
            filtered_tokens = [token for token in tokens if token.lower() not in stopword_list]
        filtered_text = ' '.join(filtered_tokens)
        return filtered_text

    # Function to get the embedding vector for n dimension, we have used "32"
    def get_embedding(self,word):
        if word in model.wv.vocab:
          return model[word]
        else:
            return np.zeros(32)
    
    # Getting average vector for each document
        out_dict = {}
        for sen in fin:
            average_vector = np.mean(np.array([self.get_embedding(x) for x in nltk.word_tokenize(self.remove_stopwords(sen))]), axis=0)
            dict = { sen : (average_vector) }
            out_dict.update(dict)

    # Function to calculate the similarity between the query vector and document vector
    def get_sim(query_embedding, average_vector_doc):
        sim = [(1 - scipy.spatial.distance.cosine(query_embedding, average_vector_doc))]
        return sim
    # Rank all the documents based on the similarity to get relevant docs
    def Ranked_documents(self,query):
        query_words = (np.mean(np.array([self.get_embedding(x) for x in nltk.word_tokenize(query.lower())],dtype=float), axis=0))
        rank = []
        for k,v in out_dict.items():
            rank.append((k, self.get_sim(query_words, v)))
        rank = sorted(rank,key=lambda t: t[1], reverse=True)[:10]
        all_obs = []
        for obs, rate in rank:
            all_obs.append(obs) 

        print('Ranked Documents :')
        print(all_obs)
        return all_obs



        
    # if __name__ == "__main__":
    #     main()`